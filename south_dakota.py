
import os
import zipfile
import json
import io
import openpyxl
import pandas as pd
from gitlab import Gitlab, GitlabError
from openpyxl.styles import PatternFill, Font
from apiClient.acsClient import acs_env_vulns
from apiClient.gitlabClient import build_scan_vunerabilites, get_all_projects, project_and_scan_componets
from apiClient.invictiClient import get_issues_for_each_scan_profile
from apiClient.jfrog import jfrog_go
from apiClient.ocpClient import ocp_env_vulns, ocp_enc_vulns
from project_csv import create_team_sheet
from utils.csv_utils import create_section_header, count_vulnerabilities, resize_column_to_max_len
from utils.file_utils import load_json_file

GITLAB_TOKEN = os.environ.get("GLAB_TOKEN")
GITLAB_URL = "https://gitlab.dillards.com"
CERT_PATH = "certs/ca.cert"

SEVERITY_COLORS = {
    "Critical": "91233e",
    "High": "de4c50",
    "Medium": "f18d43",
    "Low": "f8c851",
    "Zero": "40911d"
}


def download_sast_report(gl, projects):
    print("Downloading SAST reports...")
    for project in projects:
        if not isinstance(project, dict):
            print(f"[WARNING] Skipping invalid project: {project}")
            continue

        project_id = project.get("project_id")
        print(f"[INFO] Downloading SAST report for project {project_id}")

        try:
            glproject = gl.projects.get(project_id)
            pipelines = glproject.pipelines.list(status="success", order_by="updated_at", per_page=1)
            if not pipelines:
                print(f"[INFO] No successful pipelines for project {project_id}")
                continue

            latest_pipeline = pipelines[0]
            jobs = latest_pipeline.jobs.list()
            sast_job = next((j for j in jobs if j.name == "semgrep_sast" and j.status == "success"), None)

            if not sast_job:
                print(f"[INFO] No successful 'semgrep_sast' job for project {project_id}")
                continue

            artifact_stream = sast_job.artifacts(streamed=True)
            buf = io.BytesIO()
            for chunk in artifact_stream:
                buf.write(chunk)

            buf.seek(0)
            with zipfile.ZipFile(buf) as z:
                for file_info in z.infolist():
                    if file_info.filename.endswith(".sarif") or file_info.filename.endswith(".json"):
                        with z.open(file_info) as f:
                            sast_report = json.load(f)
                            print(f"[INFO] Loaded SAST report from {file_info.filename}")
                            return sast_report

        except GitlabError as e:
            print(f"[ERROR] GitLab error for project {project_id}: {e.error_message}")


def main():
    gl = Gitlab(url=GITLAB_URL, private_token=GITLAB_TOKEN, ssl_verify=CERT_PATH)
    all_projects = get_all_projects(gl, use_gitlab=True)
    gitlab_cicd_capable_projects = project_and_scan_componets(gl, all_projects)
    build_vulnerabilities = build_scan_vunerabilites(gitlab_cicd_capable_projects)

    ocp_dev_vulns = ocp_enc_vulns(False)
    ocp_prod_vulns = ocp_env_vulns(True)

    invicti_issues = get_issues_for_each_scan_profile()

    acs_prod_vulns = acs_env_vulns(prod=True)
    acs_dev_vulns = acs_env_vulns(prod=False)

    team_data = load_json_file("team_data.json")
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for team_name, team_info in team_data.items():
        team_urls = team_info.get("urls", [])
        team_repos = team_info.get("repos", [])
        team_groups = team_info.get("glab_group", [])

        create_team_sheet(
            workbook=workbook,
            team_name=team_name,
            team_repos=team_repos,
            team_groups=team_groups,
            team_urls=team_urls,
            build_vulnerabilities=build_vulnerabilities,
            ocp_dev_vulns=ocp_dev_vulns,
            ocp_prod_vulns=ocp_prod_vulns,
            acs_dev_vulns=acs_dev_vulns,
            acs_prod_vulns=acs_prod_vulns,
            invicti_issues=invicti_issues,
        )

    workbook.save('report.xlsx')
    print("report.xlsx created successfully!")

    download_sast_report(gl, gitlab_cicd_capable_projects)


if __name__ == "__main__":
    main()


def generate_team_sheet(sheet, start_row, team_vulns):
    headers = ["Repo", "Critical", "High", "Medium", "Low"]
    SEVERITY_COLORS = {
        "Critical": "91233e",
        "High": "de4c50",
        "Medium": "f18d43",
        "Low": "f8c851",
        "Zero": "40911d"
    }

    row = create_section_header(sheet, start_row, 1, headers)

    # Header styling
    for col_idx, col_name in enumerate(headers, 1):
        cell = sheet.cell(row=row - 1, column=col_idx)
        cell.fill = PatternFill(start_color="a8d7b2", end_color="a8d7b2", fill_type="solid")
        cell.font = Font(bold=True)

    # Data rows
    for repo_data in team_vulns:
        repo_name = repo_data["repo"]
        vulns = repo_data.get("vulnerabilities", {})
        counts = [
            repo_name,
            count_vulnerabilities(vulns, "Critical"),
            count_vulnerabilities(vulns, "High"),
            count_vulnerabilities(vulns, "Medium"),
            count_vulnerabilities(vulns, "Low"),
        ]

        for col_idx, val in enumerate(counts, 1):
            cell = sheet.cell(row=row, column=col_idx, value=val)
            if col_idx > 1:  # Only color severity columns
                severity = headers[col_idx - 1]
                color = SEVERITY_COLORS["Zero"] if val == 0 else SEVERITY_COLORS[severity]
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    resize_column_to_max_len(sheet)
    return row

def add_sast_overview_tab(wb, df):
    headers = ["Severity", "Total Count"]
    ws = wb.create_sheet("SAST Overview")
    ws.append(headers)

    for severity in ["Critical", "High", "Medium", "Low"]:
        total = df[severity].sum()
        ws.append([severity, total])
        for col_idx in range(1, 3):
            color = SEVERITY_COLORS.get(severity, "FFFFFF")
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            
 

    

def main():
    gl = Gitlab(url=GITLAB_URL, private_token=GITLAB_TOKEN, ssl_verify="certs/ca.cert")
    all_projects = get_all_projects(gl,use_gitlab=True)
    gitlab_cicd_capable_projects = project_and_scan_componets(gl,all_projects)
    build_vulnerabilities = build_scan_vunerabilites(gitlab_cicd_capable_projects)


    ocp_dev_vulns = ocp_enc_vulns(False)
    ocp_prod_vulns = ocp_env_vulns(True)

    invicti_issues = get_issues_for_each_scan_profile()

    acs_prod_vulns = acs_env_vulns(prod=True)
    acs_dev_vulns = acs_env_vulns(prod=False)

    team_data = load_json_file("team_data.json")

    workbook = openpyxl.Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for team_name, team_info in team_data.items():
        team_urls = team_info.get("urls", [])
        team_repos = team_info.get("repos", [])
        teams_group = team_info.get("glab_group", [])
        create_team_sheet(workbook=workbook, team_name=team_name, team_repos=team_repos, team_groups=team_groups,
                          team_urls = team_urls, build_vulnerabilities=build_vulnerabilities,ocp_dev_vulns=ocp_dev_vulns,
                           ocp_prod_vulns=ocp_prod_vulns, acs_dev_vulns= acs_dev_vulns, acs_prod_vulns=acs_prod_vulns, invicti_issues=invicti_issues,)
        if team_name in team_sast_vulns:
              next_row= create_merged_cell(
                    sheet,
                    row=25,
                    start_col=1,
                    end_col=5,
                    value="SAST Scan Vulnerabilities",
                    color="a8d7b2",
                    font_size=14,
              )
    
    
    
    
    
    workbook.save('report.xlsx')
    print("report.xlsx created successfully!")
    download_sast_report(gl, project_and_scan_componets(gl, all_projects))

   
if __name__ == "__main__":
    main()
#jfog_go()
#print(deployments(True))